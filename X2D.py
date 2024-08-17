import openpyxl

class X2D:
    def __init__(self):
        # Dictionaries to store mappings between countries and languages
        self.countries_by_language = {}
        self.languages_by_country = {}

    def load_data(self, countries_file, languages_file):
        # Load and process countries.xlsx
        countries_wb = openpyxl.load_workbook(countries_file)
        countries_sheet = countries_wb.active

        for row in countries_sheet.iter_rows(min_row=2, values_only=True):
            country, languages = row
            self.languages_by_country[country] = languages.split(', ')

        # Load and process languages.xlsx
        languages_wb = openpyxl.load_workbook(languages_file)
        languages_sheet = languages_wb.active

        for row in languages_sheet.iter_rows(min_row=2, values_only=True):
            language, countries = row
            self.countries_by_language[language] = countries.split(', ')
